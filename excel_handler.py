from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import json
import os
import shutil
from pathlib import Path

# === Настройки ===
LOCAL_APP_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_DIR = os.path.join(LOCAL_APP_DIR, "Шаблоны")
OUTPUT_DIR = os.path.join(LOCAL_APP_DIR, "Расчеты")
TEMPLATE_FILE = os.path.join(INPUT_DIR, "Расчет_шаблон_V1.xlsx")

def archive_existing_files(target_folder):
    """Архивирует существующие файлы в целевой папке в подпапки по порядку."""
    target_path = Path(target_folder)
    target_path.mkdir(exist_ok=True)  # Создаём целевую папку, если её нет

    # Получаем список файлов в целевой папке (без подпапок)
    files = [f for f in target_path.iterdir() if f.is_file()]

    if not files:
        print("📁 Целевая папка пуста, архивация не требуется.")
        return

    # Определяем номер следующей архивной папки
    existing_archives = [d for d in target_path.iterdir() if d.is_dir() and d.name.isdigit()]
    next_archive_num = max([int(d.name) for d in existing_archives], default=0) + 1
    archive_folder = target_path / str(next_archive_num)
    archive_folder.mkdir()

    # Переносим все файлы в архивную папку
    for file in files:
        shutil.move(str(file), str(archive_folder))

    print(f"📦 Архивная папка создана: {archive_folder.name}/")
    print(f"📊 Перемещено файлов: {len(files)}")

def export_products_to_db(deal_products, output_path, get_catalog_product_func):
    """
    Экспортирует данные о товарах из сделки в Excel файл с динамическими столбцами.
    Столбцы определяются на основе всех уникальных ключей из данных товаров (как из сделки, так и из каталога).
    Сложные типы данных (dict, list) сериализуются в строки JSON.
    :param deal_products: Список словарей, представляющих строки товаров из сделки (результат get_deal_product).
    :param output_path: Путь к Excel файлу для сохранения.
    :param get_catalog_product_func: Функция для получения данных каталога
    """

    # Список для хранения всех обработанных строк (с объединенными данными)
    all_rows_data = []
    # Множество для хранения всех уникальных ключей (названий столбцов)
    all_keys = set()

    print("Обработка товаров из сделки...")
    for product_row in deal_products:
        catalog_product_id = product_row.get("PRODUCT_ID")

        catalog_data = {}
        if catalog_product_id:
            catalog_data = get_catalog_product_func(catalog_product_id) or {}

        combined_row = {}
        # Добавляем префиксы
        for key, value in product_row.items():
            combined_row[f"DEAL_{key}"] = value

        for key, value in catalog_data.items():
            combined_row[f"CATALOG_{key}"] = value

        # Обработка значений перед добавлением в combined_row
        processed_row = {}
        for key, original_value in combined_row.items():
            # Проверяем тип значения
            if isinstance(original_value, (dict, list)):
                # Сериализуем сложный объект в строку JSON
                processed_value = json.dumps(original_value, ensure_ascii=False,
                                             indent=2)  # ensure_ascii=False для кириллицы
            else:
                # Оставляем примитивные типы как есть, или конвертируем в строку, если нужно
                # Если вы хотите все значения в виде строк, используйте str(original_value)
                # processed_value = str(original_value)
                # Но лучше оставить как есть, если это примитив, и только сложные типы сериализовать
                processed_value = original_value

            processed_row[key] = processed_value

        all_rows_data.append(processed_row)
        all_keys.update(processed_row.keys())

    headers = sorted(list(all_keys))

    print(f"Найдено уникальных полей (столбцов): {len(headers)}")
    if not headers:
        print("Предупреждение: Не найдено никаких данных для экспорта.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Товары_Сделки"

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for row_num, row_data in enumerate(all_rows_data, 2):
        for col_num, header in enumerate(headers, 1):
            value = row_data.get(header, "")  # Значение уже обработано в processed_row
            ws.cell(row=row_num, column=col_num, value=value)

    # Автоподгонка ширины столбцов (может быть медленно)
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                # Для строк, полученных из JSON, длина может быть большой
                # Рассмотрите ограничение длины, если столбцы получаются слишком широкими
                str_val = str(cell.value)
                # Ограничиваем длину для расчета ширины, если значение слишком длинное
                display_val = str_val[:50] + "..." if len(str_val) > 50 else str_val
                if len(display_val) > max_length:
                    max_length = len(display_val)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Ограничиваем максимальную ширину
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Полная выгрузка товаров успешно экспортирована в {output_path}")
    print(f"Файл содержит {len(all_rows_data)} строк товаров и {len(headers)} столбцов.")

def fill_excel(products, deal_id, get_deal_func):
    OUTPUT_FILE = os.path.join(OUTPUT_DIR, str(deal_id), f"расчет_{deal_id}.xlsx")

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)  # ← важно!
    # Шаг 3: Архивация старых файлов
    print("3️⃣ Проверяем и архивируем старые файлы...")
    archive_existing_files(os.path.join(OUTPUT_DIR, str(deal_id)))

    shutil.copy(TEMPLATE_FILE, OUTPUT_FILE)
    wb = load_workbook(OUTPUT_FILE)  # ← исправлено: было output_path (не определено)
    ws_calc = wb["Калькулятор"]
    ws_ship = wb["Доставка"]
    # === Добавляем заполнение вкладки "Сделка" ===
    print("Добавляем заполнение вкладки 'Сделка'...")
    deal = get_deal_func(deal_id)
    print(f"Полученные данные сделки: {deal}")
    
    # Заполнение вкладки "Сделка" (реализовать по необходимости)
    # ws_deal = wb["Сделка"]  # предполагаем, что лист существует
    
    # Заполнение вкладки "Калькулятор"
    start_row = 13  # Начинаем с 13-й строки
    for i, product in enumerate(products):
        row = start_row + i
        
        # Заполняем ячейки
        ws_calc[f'A{row}'] = product.get('QUANTITY', '')  # Количество
        ws_calc[f'B{row}'] = product.get('PRODUCT_NAME', '')  # Наименование
        ws_calc[f'C{row}'] = product.get('PROPERTY_234', '')  # Артикул
        ws_calc[f'D{row}'] = product.get('PROPERTY_206', '')  # Ссылка на товар
        ws_calc[f'E{row}'] = product.get('PROPERTY_216', 0)  # Цена закупа
        ws_calc[f'F{row}'] = product.get('PROPERTY_200', 0)  # Наценка
        ws_calc[f'G{row}'] = product.get('PRICE', 0)  # Цена в сделке
        ws_calc[f'H{row}'] = product.get('PROPERTY_228', '')  # Поставщик
        ws_calc[f'I{row}'] = product.get('PROPERTY_238', '')  # Срок отгрузки
        ws_calc[f'J{row}'] = product.get('PROPERTY_242', '')  # Бронируется?
        ws_calc[f'K{row}'] = product.get('PROPERTY_244', '')  # Реестр Минпрома
        ws_calc[f'L{row}'] = product.get('PROPERTY_204', '')  # Страна производства
        ws_calc[f'M{row}'] = product.get('PROPERTY_212', '')  # Реестровая запись в Минпроме
        ws_calc[f'N{row}'] = product.get('PROPERTY_214', 0)  # Объём (м³)
        ws_calc[f'O{row}'] = product.get('PROPERTY_232', 0)  # Вес (г)
        ws_calc[f'P{row}'] = product.get('PROPERTY_194', '')  # Техническое описание

    # Заполнение вкладки "Доставка"
    # (реализовать по необходимости)
    
    wb.save(OUTPUT_FILE)
    print(f"Excel файл заполнен: {OUTPUT_FILE}")