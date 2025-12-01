import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import json  # Импортируем json для сериализации сложных объектов
import os
import shutil

from pathlib import Path

import logging
import builtins

from config import OUTPUT_DIR, TEMPLATE_FILE, DEAL_ID, COLUMN_LABELS
from key import WEBHOOK_URL

# Включить логирование: DEBUG, INFO, WARNING...
# Отключить: logging.CRITICAL + 1
#LOG_LEVEL = logging.CRITICAL + 1  # отключить
LOG_LEVEL =  logging.INFO
logging.basicConfig(level=LOG_LEVEL, format='%(message)s')

def fast_print(*args, **kwargs):
    if logging.root.level <= logging.INFO:
        message = ' '.join(str(x) for x in args)
        logging.info(message)

builtins.print = fast_print

def update():
    from updater import update
    update()




def get_deal(deal_id):
    """Получает общие данные из сделки"""
    url = f"{WEBHOOK_URL}crm.deal.get"
    response = requests.post(url, json={"id": DEAL_ID})
    response.raise_for_status()
    return response.json().get("result", {})


def get_catalog_element(element_id):
    """Получает детальные данные товара из каталога"""
    url = f"{WEBHOOK_URL}crm.product.get"
    try:
        response = requests.post(url, json={"id": element_id})
        response.raise_for_status()
        return response.json().get("result", {})
    except requests.exceptions.HTTPError as e:
        if response.status_code == 400:
            print(f"Товар с ID {element_id} не найден в каталоге (ручная позиция)")
            return None  # Возвращаем None для ручных позиций
        else:
            raise e


def archive_existing_files(target_folder):
    """Архивирует существующие файлы в целевой папке в подпапки по порядку."""
    target_path = Path(target_folder)
    target_path.mkdir(exist_ok=True)  # Создаём целевую папку, если её нет

    # Получаем список файлов в целевой папке (без подпапок)
    files = [f for f in target_path.iterdir() if f.is_file()]

    if not files:
        print("📁 Целевая папка пуста, архивация не требуется.")
        return

    # Определяем номер следующей архивной папки
    existing_archives = [d for d in target_path.iterdir() if d.is_dir() and d.name.isdigit()]
    next_archive_num = max([int(d.name) for d in existing_archives], default=0) + 1
    archive_folder = target_path / str(next_archive_num)
    archive_folder.mkdir()

    # Переносим все файлы в архивную папку
    for file in files:
        shutil.move(str(file), str(archive_folder))

    print(f"📦 Архивная папка создана: {archive_folder.name}/")
    print(f"📊 Перемещено файлов: {len(files)}")




# === 1. Получение строк товаров из сделки ===
def get_deal_products(deal_id):
    """Получает товары из сделки"""
    url = f"{WEBHOOK_URL}crm.deal.productrows.get"
    response = requests.post(url, json={"id": deal_id})
    response.raise_for_status()
    return response.json().get("result", [])


# === 2. Получение данных товара из каталога магазина ===
def get_catalog_product(product_id):
    url = f"{WEBHOOK_URL}catalog.product.get"
    response = requests.post(url, json={"id": product_id})
    if response.status_code == 200:
        return response.json().get("result", {})
    else:
        print(f"⚠️ Не удалось загрузить товар ID={product_id}")
        return None


# === 3. Полная выгрузка товаров в Excel с динамическими полями и обработкой сложных типов ===
def export_products_to_db(deal_products, output_path):
    """
    Экспортирует данные о товарах из сделки в Excel файл с динамическими столбцами.
    Столбцы определяются на основе всех уникальных ключей из данных товаров (как из сделки, так и из каталога).
    Сложные типы данных (dict, list) сериализуются в строки JSON.
    :param deal_products: Список словарей, представляющих строки товаров из сделки (результат get_deal_product).
    :param output_path: Путь к Excel файлу для сохранения.
    """

    # Список для хранения всех обработанных строк (с объединенными данными)
    all_rows_data = []
    # Множество для хранения всех уникальных ключей (названий столбцов)
    all_keys = set()

    print("Обработка товаров из сделки...")
    for product_row in deal_products:
        catalog_product_id = product_row.get("PRODUCT_ID")

        catalog_data = {}
        if catalog_product_id:
            catalog_data = get_catalog_product(catalog_product_id) or {}

        combined_row = {}
        # Добавляем префиксы
        for key, value in product_row.items():
            combined_row[f"DEAL_{key}"] = value

        for key, value in catalog_data.items():
            combined_row[f"CATALOG_{key}"] = value

        # Обработка значений перед добавлением в combined_row
        processed_row = {}
        for key, original_value in combined_row.items():
            # Проверяем тип значения
            if isinstance(original_value, (dict, list)):
                # Сериализуем сложный объект в строку JSON
                processed_value = json.dumps(original_value, ensure_ascii=False,
                                             indent=2)  # ensure_ascii=False для кириллицы
            else:
                # Оставляем примитивные типы как есть, или конвертируем в строку, если нужно
                # Если вы хотите все значения в виде строк, используйте str(original_value)
                # processed_value = str(original_value)
                # Но лучше оставить как есть, если это примитив, и только сложные типы сериализовать
                processed_value = original_value

            processed_row[key] = processed_value

        all_rows_data.append(processed_row)
        all_keys.update(processed_row.keys())

    headers = sorted(list(all_keys))

    print(f"Найдено уникальных полей (столбцов): {len(headers)}")
    if not headers:
        print("Предупреждение: Не найдено никаких данных для экспорта.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Товары_Сделки"

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for row_num, row_data in enumerate(all_rows_data, 2):
        for col_num, header in enumerate(headers, 1):
            value = row_data.get(header, "")  # Значение уже обработано в processed_row
            ws.cell(row=row_num, column=col_num, value=value)

    # Автоподгонка ширины столбцов (может быть медленно)
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                # Для строк, полученных из JSON, длина может быть большой
                # Рассмотрите ограничение длины, если столбцы получаются слишком широкими
                str_val = str(cell.value)
                # Ограничиваем длину для расчета ширины, если значение слишком длинное
                display_val = str_val[:50] + "..." if len(str_val) > 50 else str_val
                if len(display_val) > max_length:
                    max_length = len(display_val)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Ограничиваем максимальную ширину
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Полная выгрузка товаров успешно экспортирована в {output_path}")
    print(f"Файл содержит {len(all_rows_data)} строк товаров и {len(headers)} столбцов.")


def fill_excel(products, deal_id):
    OUTPUT_FILE = os.path.join(OUTPUT_DIR, str(deal_id), f"расчет_{deal_id}.xlsx")

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)  # ← важно!
    # Шаг 3: Архивация старых файлов
    print("3️⃣ Проверяем и архивируем старые файлы...")
    archive_existing_files(os.path.join(OUTPUT_DIR, str(deal_id)))

    shutil.copy(TEMPLATE_FILE, OUTPUT_FILE)
    wb = load_workbook(OUTPUT_FILE)  # ← исправлено: было output_path (не определено)
    ws_calc = wb["Калькулятор"]
    ws_ship = wb["Доставка"]
    # === Добавляем заполнение вкладки "Сделка" ===
    print("Добавляем заполнение вкладки 'Сделка'...")
    deal = get_deal(deal_id)
    print(f"Полученные данные сделки: {list(deal.keys())}")  # логирование

    if "Сделка" in wb.sheetnames:
        ws_deal = wb["Сделка"]
        print("Найдена вкладка 'Сделка', начинаем заполнение...")
        fill_deal_sheet(ws_deal, deal, deal_id)
        print("Заполнение вкладки 'Сделка' завершено")
    else:
        print("Вкладка 'Сделка' не найдена в шаблоне")
        print(f"Доступные вкладки: {wb.sheetnames}")

    # === Запись товаров на вкладку "Товары" с двухстрочной шапкой ===
    df_products = products_to_excel(deal_id)  # ← см. ниже: функция без переименования колонок!

    if "Товары" in wb.sheetnames:
        ws_prod = wb["Товары"]
    else:
        ws_prod = wb.create_sheet("Товары")

    # Получаем технические названия колонок (в порядке COLUMN_LABELS)
    technical_columns = list(COLUMN_LABELS.keys())

    # 1️⃣ Строка 1: технические названия
    for col_idx, col_name in enumerate(technical_columns, start=1):
        ws_prod.cell(row=1, column=col_idx, value=col_name)

    # 2️⃣ Строка 2: русские метки
    for col_idx, col_name in enumerate(technical_columns, start=1):
        ws_prod.cell(row=2, column=col_idx, value=COLUMN_LABELS[col_name])

    # 3️⃣ Строки 3+: данные
    for row_idx, row_data in enumerate(df_products.itertuples(index=False, name=None), start=3):
        for col_idx, value in enumerate(row_data, start=1):
            ws_prod.cell(row=row_idx, column=col_idx, value=value)
    # === 1. Лист "Доставка" — заполняем только входные данные ===
    # Очищаем только входные столбцы (A–H), формулы в I–K останутся
    ##    for row in range(3, 98):
    ##        for col in "ABCDEFGH":
    ##            ws_ship[f"{col}{row}"].value = None
    start_row = 3
    max_rows = 198

    #for i, p in enumerate(products):
        #r = i + 3
        # ws_ship[f"B{r}"] = p["name"]
        # ws_ship[f"C{r}"] = p["quantity"]
        # ws_ship[f"D{r}"] = p["supplier"]
        #ws_ship[f"E{r}"] = p["shipping_city"]
        #ws_ship[f"F{r}"] = float(p["weight_g"])
        #ws_ship[f"G{r}"] = p["length_mm"]
        #ws_ship[f"H{r}"] = p["width_mm"]
        #ws_ship[f"I{r}"] = p["height_mm"]
        #if p["volume_m3"]:
            #ws_ship[f"J{r}"] = float(p["volume_m3"])

    # ws_ship[f"F1"] = p["height_mm"]

    # === Скрываем пустые строки с товарами ===
    num_products = len(products)
    for i in range(num_products, max_rows):
        r = start_row + i
        ws_ship.row_dimensions[r].hidden = True

    # Строки "Доставка" (199) и "Итого" (200) — всегда видимы
    ws_ship.row_dimensions[199].hidden = False
    ws_ship.row_dimensions[200].hidden = False

    # === 2. Лист "Калькулятор" — заполняем ТОЛЬКО входные ячейки ===
    start_row = 3
    max_rows = 198  # строки 3–42

    # Очищаем только входные столбцы (B–G), остальное — формулы!
    # for i in range(max_rows):
    #     r = start_row + i
    #     for col in "BCDEFG":
    #         ws_calc[f"{col}{r}"].value = None

    # # Заполняем реальные товары
    # for i, p in enumerate(products):
    #     r = start_row + i
    #     ws_calc[f"B{r}"] = p["product_id"]
    #     ws_calc[f"C{r}"] = p["name"]
    #     ws_calc[f"D{r}"] = p["quantity"]
    #     ws_calc[f"E{r}"] = float(p["price_purchase"])
    #     ws_calc[f"F{r}"] = p["supplier"]
    #     ws_calc[f"G{r}"] = p["vat_in"]
    #     ws_calc[f"H{r}"] = p["link"]
    #     ws_calc[f"Z{r}"] = p["bron"]
    #     ws_calc[f"Y{r}"] = p["actual"]

    # Строка "Доставка" (99) — заполняем только входные поля
    # ws_calc[f"С99"] = "Доставка"
    # ws_calc[f"D99"] = 1
    # ws_calc[f"D99"] = 27900  # можно параметризовать
    # ws_calc[f"E99"] = "СДЭК"
    # ws_calc[f"F99"] = "УСН"
    # ws_calc[f"G99"] = ""

    # Строка "Итого" (100) — только текст, формулы уже есть
    # ws_calc[f"B100"] = "Итого"

    # === Скрываем пустые строки с товарами ===
    num_products = len(products)
    for i in range(num_products, max_rows):
        r = start_row + i
        ws_calc.row_dimensions[r].hidden = True

    # Строки "Доставка" (99) и "Итого" (100) — всегда видимы
    ws_calc.row_dimensions[199].hidden = False
    ws_calc.row_dimensions[200].hidden = False

    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE


def import_data(deal_id):
    print("1️⃣ Получаем строки товаров из сделки...")
    rows = get_deal_products(deal_id)
    if not rows:
        print("❌ В сделке нет товаров.")
        OUTPUT_FILE = os.path.join(OUTPUT_DIR, str(deal_id), f"расчет_{deal_id}.xlsx")
        os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
        shutil.copy(TEMPLATE_FILE, OUTPUT_FILE)
        print(f"📁 Создан пустой файл из шаблона: {OUTPUT_FILE}")
        return

    print("2️⃣ Собираем уникальные product_id для пакетной загрузки...")
    product_ids = {row.get("PRODUCT_ID") for row in rows if row.get("PRODUCT_ID")}

    # Пакетная загрузка каталога — предполагаем, что get_catalog_products принимает список ID
    catalog_cache = {}
    if product_ids:
        try:
            # ⚠️ НУЖНО реализовать или адаптировать get_catalog_products!
            catalog_cache = get_catalog_products(list(product_ids))  # возвращает {id: data}
        except Exception as e:
            print(f"⚠️ Ошибка при загрузке каталога: {e}")
            catalog_cache = {}

    print("3️⃣ Обрабатываем строки товаров...")
    products_for_excel = []

    for row in rows:
        product_id = row.get("PRODUCT_ID")
        name = row.get("PRODUCT_NAME", "").strip()
        quantity = row.get("QUANTITY", 1)
        price_with_vat = row.get("PRICE", 0)

        if not name:
            name = f"Товар ID={product_id}" if product_id else "Неизвестный товар"

        # Получаем данные из кэша
        product = {}
        if product_id and product_id in catalog_cache:
            product = catalog_cache[product_id].get("product", {})

        # Извлечение данных с безопасными дефолтами
        weight_g = product.get("weight") or (product.get("property232") or {}).get("value", 0)
        length_mm = product.get("length", 0)
        width_mm = product.get("width", 0)
        height_mm = product.get("height", 0)
        supplier = (product.get("property228") or {}).get("value", "Не указан")
        vat_in = (product.get("property236") or {}).get("value", "НДС не указан")
        link = (product.get("property206") or {}).get("value", "")
        price_purchase = (product.get("property216") or {}).get("value", 0)
        volume_m3 = (product.get("property214") or {}).get("value", 0)
        bron = (product.get("property242") or {}).get("value", 0)
        actual = (product.get("property240") or {}).get("value", "")

        products_for_excel.append({
            "product_id": product_id,
            "name": name,
            "quantity": quantity,
            "price_with_vat": price_with_vat,
            "supplier": supplier,
            "shipping_city": "Москва",
            "weight_g": weight_g,
            "length_mm": length_mm,
            "width_mm": width_mm,
            "height_mm": height_mm,
            "vat_in": vat_in,
            "link": link,
            "price_purchase": price_purchase,
            "volume_m3": volume_m3,
            "bron": bron,
            "actual": actual
        })

    print("4️⃣ Заполняем Excel...")
    output_file = fill_excel(products_for_excel, deal_id)
    print(f"✅ Готово! Файл сохранён: {output_file}")

def get_catalog_products(ids):
    result = {}
    for pid in ids:
        result[pid] = get_catalog_product(pid)  # старая функция
    return result

def create_products_from_tovary(deal_id, webhook_url=WEBHOOK_URL, output_dir=OUTPUT_DIR):
    """
    Создаёт в каталоге товары для строк листа 'Товары', у которых PRODUCT_ID пустой или равен 0.
    Если NAME пустое, но есть PRODUCT_NAME — подставляет его в NAME.
    После создания записывает новый PRODUCT_ID обратно в Excel.
    """
    input_file = os.path.join(output_dir, str(deal_id), f"расчет_{deal_id}.xlsx")
    if not os.path.exists(input_file):
        print(f"❌ Файл не найден: {input_file}")
        return

    wb = load_workbook(input_file)
    if "Товары" not in wb.sheetnames:
        print("❌ Лист 'Товары' не найден в Excel")
        return

    ws = wb["Товары"]

    # Читаем технические имена колонок из первой строки
    headers = []
    col = 1
    while True:
        val = ws.cell(row=1, column=col).value
        if not val:
            break
        headers.append(str(val).strip())
        col += 1

    # Определяем индекс колонки PRODUCT_ID
    try:
        product_id_col = headers.index("PRODUCT_ID") + 1
    except ValueError:
        print("❌ В листе 'Товары' нет колонки PRODUCT_ID")
        return

    row_idx = 3  # начинаем с 3-й строки, т.к. 2-я строка — русские подписи
    created_count = 0

    while True:
        row_data = {}
        empty = True
        for col_idx, key in enumerate(headers, start=1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val not in (None, ""):
                empty = False
            row_data[key] = val
        if empty:
            break

        product_id = row_data.get("PRODUCT_ID")
        name = str(row_data.get("NAME") or "").strip()
        product_name = str(row_data.get("PRODUCT_NAME") or "").strip()

        # # Автозамена: если NAME пустое, но есть PRODUCT_NAME → используем его
        # if not name and product_name:
        #     name = product_name

        # Проверка: пустой или 0 любого типа
        is_manual = False
        if product_id is None or str(product_id).strip() == "":
            is_manual = True
        else:
            try:
                if float(product_id) == 0:
                    is_manual = True
            except:
                pass

        if is_manual and name:
            fields = {"NAME": name}

            if row_data.get("PRICE") not in (None, ""):
                try:
                    fields["PRICE"] = float(row_data["PRICE"])
                except:
                    pass

            if row_data.get("CURRENCY_ID"):
                fields["CURRENCY_ID"] = str(row_data["CURRENCY_ID"])

            if row_data.get("VAT_INCLUDED") in ("Y", "N"):
                fields["VAT_INCLUDED"] = row_data["VAT_INCLUDED"]

            # Все PROPERTY_XXX → {"value": ...}
            for key, value in row_data.items():
                if key.startswith("PROPERTY_") and value not in (None, ""):
                    fields[key] = {"value": value}

            url = f"{webhook_url}crm.product.add"
            payload = {"fields": fields}
            try:
                resp = requests.post(url, json=payload)
                resp.raise_for_status()
                result = resp.json()
                new_id = result.get("result")
                if new_id:
                    ws.cell(row=row_idx, column=product_id_col, value=new_id)
                    created_count += 1
                    print(f"✅ Создан товар ID={new_id} для строки {row_idx} (\"{name}\")")
                else:
                    print(f"⚠️ Не удалось создать товар для строки {row_idx}: {result}")
            except Exception as e:
                print(f"❌ Ошибка при создании товара для строки {row_idx}: {e}")
                if 'resp' in locals():
                    print(resp.text)

        row_idx += 1

    if created_count > 0:
        wb.save(input_file)
        print(f"💾 Excel обновлён: записаны новые PRODUCT_ID ({created_count} товаров).")
    else:
        print("ℹ️ Новых товаров для создания не найдено.")


# === Вспомогательные: считывание листа "Калькулятор" ===
def _read_products_from_calculator(wb):
    """
    Читает товары с листа 'Калькулятор' (строки 3–42).
    Возвращает список словарей с ключами:
    - product_id (int или None)
    - name (str)
    - quantity (int)
    - price_unit (float) — цена за единицу из колонки V (22)
    """
    ws_calc = wb["Калькулятор"]
    rows = []
    for row in range(3, 198):
        product_id_raw = ws_calc.cell(row=row, column=2).value  # B
        name = str(ws_calc.cell(row=row, column=3).value or "").strip()  # C
        qty_raw = ws_calc.cell(row=row, column=4).value  # D
        price_unit_raw = ws_calc.cell(row=row, column=22).value  # V — цена за штуку (ЛШО)

        # Пропускаем служебные строки
        if name.lower() in ("доставка", "итого"):
            continue

        # Стоп на полностью пустых строках
        if product_id_raw in (None, "") and not name and qty_raw in (None, "") and price_unit_raw in (None, ""):
            break

        # Приведение типов
        try:
            product_id = int(float(product_id_raw)) if product_id_raw not in (None, "") else None
        except (ValueError, TypeError):
            product_id = None

        try:
            quantity = int(float(qty_raw)) if qty_raw not in (None, "") else 1
        except (ValueError, TypeError):
            quantity = 1

        try:
            price_unit = float(price_unit_raw) if price_unit_raw not in (None, "") else 0.0
        except (ValueError, TypeError):
            price_unit = 0.0

        rows.append({
            "product_id": product_id,
            "name": name,
            "quantity": quantity,
            "price_unit": price_unit,
        })
    return rows


# === Вспомогательные: считывание листа "Товары" ===
def _read_products_from_products_sheet(wb):
    """
    Читает лист 'Товары' c двухстрочной шапкой:
    - Строка 1: технические имена (например, PRODUCT_ID, PROPERTY_216)
    - Строка 2: русские метки
    Возвращает список словарей, где ключи — технические имена из строки 1.
    """
    if "Товары" not in wb.sheetnames:
        return []

    ws = wb["Товары"]

    # Собираем технические имена из первой строки
    technical_headers = []
    col = 1
    while True:
        val = ws.cell(row=1, column=col).value
        if val is None:
            break
        technical_headers.append(str(val).strip())
        col += 1

    if not technical_headers:
        return []

    # Считываем строки данных, начиная с 3-й
    rows = []
    row_idx = 3
    while True:
        # Если вся строка пустая — завершаем
        empty = True
        row_data = {}
        for col_idx, tech_name in enumerate(technical_headers, start=1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val not in (None, ""):
                empty = False
            row_data[tech_name] = val
        if empty:
            break
        rows.append(row_data)
        row_idx += 1

    return rows


# === Bitrix: обновление строк сделки ===
def _bitrix_set_rows(deal_id, rows, webhook_url=WEBHOOK_URL):
    """
    Устанавливает строки сделки (полная замена).
    Ожидаемый формат rows: список словарей Bitrix crm.deal.productrows.set
    """
    url = f"{webhook_url}crm.deal.productrows.set"
    payload = {"id": deal_id, "rows": rows}
    resp = requests.post(url, json=payload)
    try:
        resp.raise_for_status()
        print(f"✅ Установлено {len(rows)} строк в сделке {deal_id}.")
    except Exception as e:
        print(f"❌ Ошибка установки строк сделки: {e}")
        print("Ответ Bitrix24:", getattr(resp, "text", ""))


# === Bitrix: обновление товара каталога ===
def _bitrix_update_product(product_id, fields, webhook_url=WEBHOOK_URL):
    """
    Обновляет товар каталога crm.product.update (или catalog.product.update в зависимости от портала).
    По твоей логике в этом файле используется crm.product.update.
    fields: словарь полей (например, NAME, PRICE, VAT_INCLUDED, CURRENCY_ID)
    """
    url = f"{webhook_url}crm.product.update"
    payload = {"id": product_id, "fields": fields}
    resp = requests.post(url, json=payload)
    try:
        resp.raise_for_status()
        print(f"✅ Товар каталога {product_id} обновлён: {fields}")
    except Exception as e:
        print(f"❌ Ошибка обновления товара {product_id}: {e}")
        print("Ответ Bitrix24:", getattr(resp, "text", ""))


def export_data(deal_id, webhook_url=WEBHOOK_URL, output_dir=OUTPUT_DIR, update_catalog=True):
    input_file = os.path.join(output_dir, str(deal_id), f"расчет_{deal_id}.xlsx")
    if not os.path.exists(input_file):
        print(f"❌ Файл не найден: {input_file}")
        return

    # === ШАГ 1: Создаём новые товары И СРАЗУ ОБНОВЛЯЕМ EXCEL ===
    print("🆕 Создаём новые товары из листа 'Товары'...")
    create_products_from_tovary(deal_id)  # ← эта функция ДОЛЖНА обновить Excel

    # === ШАГ 2: Перезагружаем файл, чтобы учесть новые PRODUCT_ID ===
    wb = load_workbook(input_file, data_only=True)

    # === ШАГ 3: Формируем строки сделки с учётом новых ID ===
    deal_rows_payload = []
    ws_calc = wb["Калькулятор"]
    row_idx = 3
    while True:
        product_id_raw = ws_calc.cell(row=row_idx, column=2).value  # B
        name = str(ws_calc.cell(row=row_idx, column=3).value or "").strip()  # C
        qty_raw = ws_calc.cell(row=row_idx, column=4).value  # D
        price_unit_raw = ws_calc.cell(row=row_idx, column=21).value  # U

        if not name or name.lower() in ("доставка", "итого"):
            if product_id_raw in (None, "") and not name and qty_raw in (None, ""):
                break
            row_idx += 1
            continue

        # Приведение ID: если 0 → считаем ручным товаром
        try:
            product_id = int(float(product_id_raw)) if product_id_raw not in (None, "") else 0
        except:
            product_id = 0

        # Если ID == 0 → это ручной товар, отправляем по имени
        row_payload = {
            "QUANTITY": int(float(qty_raw)) if qty_raw not in (None, "") else 1,
            "PRICE": float(price_unit_raw) if price_unit_raw not in (None, "") else 0.0,
            "TAX_RATE": "20.00",
            "TAX_INCLUDED": "Y",
            "CUSTOMIZED": "Y",
        }

        if product_id != 0:
            row_payload["PRODUCT_ID"] = product_id
        else:
            row_payload["PRODUCT_NAME"] = name or "Без названия"

        deal_rows_payload.append(row_payload)
        row_idx += 1

    # === ШАГ 4: Отправляем строки сделки ===
    if deal_rows_payload:
        print(f"📤 Отправляем {len(deal_rows_payload)} строк в сделку {deal_id}...")
        try:
            url = f"{webhook_url}crm.deal.productrows.set"
            resp = requests.post(url, json={"id": deal_id, "rows": deal_rows_payload})
            resp.raise_for_status()
            print("✅ Строки сделки успешно обновлены.")
        except Exception as e:
            print(f"❌ Ошибка при обновлении строк сделки: {e}")
            if 'resp' in locals():
                print(resp.text)
    else:
        print("ℹ️ Нет строк для обновления сделки.")

    # === ШАГ 5: Обновляем каталог (только для строк с PRODUCT_ID != 0) ===
    if update_catalog:
        print("🔄 Обновляем товары каталога из листа 'Товары'...")
        if "Товары" not in wb.sheetnames:
            print("❌ Лист 'Товары' не найден.")
            return

        ws_prod = wb["Товары"]
        # Читаем технические заголовки из строки 1
        headers = []
        col = 1
        while True:
            val = ws_prod.cell(row=1, column=col).value
            if val is None:
                break
            headers.append(str(val).strip())
            col += 1

        if not headers:
            print("❌ Не найдены заголовки на листе 'Товары'.")
            return

        updated_count = 0
        row_idx = 3
        while True:
            row_data = {}
            empty = True
            for col_idx, key in enumerate(headers, start=1):
                val = ws_prod.cell(row=row_idx, column=col_idx).value
                if val not in (None, ""):
                    empty = False
                    row_data[key] = val

            # Прерываем по пустому PRODUCT_NAME
            if row_data.get("PRODUCT_NAME") in (None, ""):
                break
            if empty:
                break

            product_id_raw = row_data.get("PRODUCT_ID")
            if product_id_raw is None or str(product_id_raw).strip() == "":
                is_manual = True
            else:
                try:
                    is_manual = float(product_id_raw) == 0
                except:
                    is_manual = True

            if is_manual:
                row_idx += 1
                continue

            product_id = int(float(product_id_raw))
            fields = {}

            # NAME
            name_val = row_data.get("PRODUCT_NAME")
            if name_val:
                fields["NAME"] = str(name_val).strip()

            # PRICE
            price_val = row_data.get("PRICE")
            if price_val is not None:
                try:
                    fields["PRICE"] = float(price_val)
                except:
                    pass

            # CURRENCY_ID, VAT_INCLUDED
            if row_data.get("CURRENCY_ID"):
                fields["CURRENCY_ID"] = str(row_data["CURRENCY_ID"])
            if row_data.get("VAT_INCLUDED") in ("Y", "N"):
                fields["VAT_INCLUDED"] = row_data["VAT_INCLUDED"]

            # PROPERTY_XXX
            for key in headers:
                if key.startswith("PROPERTY_"):
                    if key == "PROPERTY_202":
                        from datetime import datetime
                        current_date_iso = datetime.now().strftime('%Y-%m-%dT%H:%M:%S+03:00')
                        fields[key] = {"value": current_date_iso}
                    elif key == "PROPERTY_240":
                        fields[key] = {"value": "Да"}
                    else:
                        val = row_data.get(key)
                        if val not in (None, ""):
                            fields[key] = {"value": val}

            # Отправка обновления
            try:
                resp = requests.post(
                    f"{webhook_url}crm.product.update",
                    json={"id": product_id, "fields": fields}
                )
                resp.raise_for_status()
                updated_count += 1
                print(f"✅ Товар {product_id} обновлён.")
            except Exception as e:
                print(f"❌ Ошибка обновления товара {product_id}: {e}")

            row_idx += 1

        print(f"🎯 Каталог обновлён: {updated_count} товаров.")

        
def export_data_КЕДО(deal_id):
    print("\n📤 ЭКСПОРТ КЕДО (цены из колонки x = 24)")
    _export_data_with_price_column(
        deal_id,
        price_col_index=24,  # колонка х
        tax_rate="0.00",  # Без НДС
        tax_included="N",  # НДС не включён в цену
        supplier_name="ИП",
        mode="КЕДО"
    )


def export_data_Verch(deal_id):
    print("\n📤 ЭКСПОРТ ВЕРШ (цены из колонки W = 23)")
    _export_data_with_price_column(
        deal_id,
        price_col_index=23,  # колонка T
        tax_rate="0.00",  # Без НДС
        tax_included="N",  # НДС не включён
        supplier_name="ИП",
        mode="Верш"
    )


def export_data_LSO(deal_id):
    print("\n📤 ЭКСПОРТ ЛШО (цены из колонки V = 22)")
    _export_data_with_price_column(
        deal_id,
        price_col_index=22,  # колонка Q — "Цена ЛШО"
        tax_rate="20.00",  # НДС 20%
        tax_included="Y",  # НДС включён в цену
        supplier_name="ИП",
        mode="ЛШО"
    )


def _export_data_with_price_column(deal_id, price_col_index, tax_rate, tax_included, supplier_name, mode):
    input_file = os.path.join(OUTPUT_DIR, str(deal_id), f"расчет_{deal_id}.xlsx")
    if not os.path.exists(input_file):
        print(f"❌ Файл не найден: {input_file}")
        return

    wb = load_workbook(input_file, data_only=True)
    ws = wb["Калькулятор"]

    new_rows = []
    for row in range(3, 196):
        product_id_raw = ws.cell(row=row, column=2).value  # B
        name = str(ws.cell(row=row, column=3).value or "").strip()

        if not name or name.lower() in ("доставка", "итого"):
            if product_id_raw in (None, "") and not name:
                break
            # "Доставка" и "Итого" не добавляем как товары

        try:
            product_id = int(float(product_id_raw)) if product_id_raw not in (None, "") else 0
        except (ValueError, TypeError):
            product_id = 0

        quantity_raw = ws.cell(row=row, column=4).value  # D
        price_raw = ws.cell(row=row, column=price_col_index).value  # нужная колонка

        quantity = int(float(quantity_raw)) if quantity_raw not in (None, "") else 1
        price = float(price_raw) if price_raw not in (None, "") else 0.0

        row_data = {
            "QUANTITY": quantity,
            "PRICE": price,
            "TAX_RATE": tax_rate,
            "TAX_INCLUDED": tax_included,
        }

        if product_id == 0:
            if name:
                row_data["PRODUCT_NAME"] = name
                new_rows.append(row_data)
        else:
            row_data["PRODUCT_ID"] = product_id
            new_rows.append(row_data)

        #print(f"   ✅ {name} → {price} руб, НДС: {tax_rate}, Включён: {tax_included}")

    # Отправка в Bitrix24
    try:
        response = requests.post(
            f"{WEBHOOK_URL}crm.deal.productrows.set",
            json={"id": deal_id, "rows": new_rows}
        )
        response.raise_for_status()
        #print(f"✅ Успешно обновлено ({mode})!")
    except Exception as e:
        print(f"❌ Ошибка при экспорте {mode}: {e}")
        if 'response' in locals():
            print(f"Ответ Bitrix24:", json.dumps(response.text))


def generate_3kp(deal_id):
    """
    Генерация 3 КП:
    - КЕДО (цены из колонки U)
    - Верш (цены из колонки T)
    """
    print(f"🖨️ Генерация 3 КП для сделки {deal_id}...")

    # Экспорт для КЕДО
    print("\n➡️ Экспорт КЕДО...")
    export_data_КЕДО(deal_id)
    generate_KP(deal_id, 50)
    print("\n➡️ Экспорт Верхозина...")
    export_data_Verch(deal_id)
    generate_KP(deal_id, 48)
    print("\n➡️ Экспорт ЛШО...")
    export_data_LSO(deal_id)
    generate_KP(deal_id, 46)

    print("\n✅ Генерация 3 КП завершена.")

def generate_kp_lsho(deal_id):
    export_data_LSO(deal_id)
    generate_KP(deal_id, 46)

def generate_kp_verch(deal_id):
    export_data_Verch(deal_id)
    generate_KP(deal_id, 48)


def generate_KP(deal_id, template_id=46, webhook_url=WEBHOOK_URL, entity_type_id=2, output_dir="Расчеты"):
    """
    Вызывает метод crm.documentgenerator.document.add через вебхук,
    имитируя BX24.callMethod, и сохраняет сгенерированный документ в указанный каталог.

    Args:
        entity_id (str or int): ID сущности (например, сделки).
        template_id (int): ID шаблона:
            46 → "КП ЛШО",
            48 → "КП Верхозина",
            50 → "КП КЕДО",
            52 → "КП ЛШО с фото".
        webhook_url (str): URL вебхука Bitrix24.
        entity_type_id (int or str): Тип сущности (2 = сделка).
        output_dir (str): Директория для сохранения файлов.

    Returns:
        dict: Результат API + путь к сохранённому файлу (если успешно).
    """
    # === 1. Получаем значение из Excel (вкладка "Сделка", ячейка C26) ===
    input_file = os.path.join(OUTPUT_DIR, str(deal_id), f"расчет_{deal_id}.xlsx")
    try:
        wb = load_workbook(input_file, data_only=True)
        ws_deal = wb["Сделка"]
        delivery_term_cell = ws_deal['C26'].value
        # Если значение пустое или 0, устанавливаем 3
        if not delivery_term_cell or delivery_term_cell == 0:
            delivery_term_value = 3
        else:
            delivery_term_value = int(delivery_term_cell)  # Преобразуем в целое число, если возможно
    except Exception as e:
        print(f"⚠️ Не удалось прочитать срок поставки из Excel: {e}")
        delivery_term_value = 3  # По умолчанию

    # === 2. Обновляем сделку в Bitrix24 с новым значением ===
    update_payload = {
        "id": deal_id,
        "fields": {
            "UF_CRM_1757398927169": delivery_term_value
        }
    }
    try:
        update_url = f"{webhook_url.rstrip('/')}/crm.deal.update.json"
        update_resp = requests.post(update_url, json=update_payload)
        update_resp.raise_for_status()
        print(f"✅ Успешно обновлено пользовательское поле UF_CRM_1757398927169 со значением {delivery_term_value}")
    except Exception as e:
        print(f"❌ Ошибка при обновлении сделки: {e}")

    # Определяем префикс имени файла по template_id
    template_names = {
        46: "КП ЛШО",
        48: "КП Верхозина",
        50: "КП КЕДО",
        52: "КП ЛШО с фото"
    }
    template_name = template_names.get(template_id, f"КП_шаблон_{template_id}")

    # Формируем путь к файлу

    full_path = os.path.join(OUTPUT_DIR, str(deal_id), f"{template_name}_{deal_id}.docx")

    def download_document(download_url, save_path):
        """Загружает документ и сохраняет по указанному пути."""
        if not download_url:
            print("❌ Отсутствует downloadUrl для загрузки документа.")
            return None

        try:
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            print(f"📥 Загрузка документа по URL: {download_url}")
            resp = requests.get(download_url)
            resp.raise_for_status()

            with open(save_path, 'wb') as f:
                f.write(resp.content)
            print(f"✅ Документ успешно сохранён: {save_path}")
            return save_path

        except Exception as e:
            print(f"❌ Ошибка при загрузке документа: {e}")
            return None

    # --- Основной вызов API ---
    api_method = 'crm.documentgenerator.document.add'
    url = f"{webhook_url.rstrip('/')}/{api_method}.json"

    payload = {
        'templateId': template_id,
        'entityTypeId': entity_type_id,
        'entityId': deal_id
    }

    print(f"Попытка вызвать метод {api_method} с параметрами: {payload}")
    print(f"URL запроса: {url}")

    try:
        response = requests.post(url, json=payload)
        response.raise_for_status()
        result = response.json()



        if 'result' in result and 'document' in result['result']:
            download_url = result['result']['document'].get('downloadUrlMachine')
            if download_url:
                saved_path = download_document(download_url, full_path)
                if saved_path:
                    result['downloaded_file'] = saved_path
            else:
                print("⚠️ В ответе отсутствует downloadUrl.")
            print("Документ успешно создан.")
            return result
        else:
            print(f"⚠️ Некорректный ответ API: {result}")
            return result

    except requests.exceptions.HTTPError as e:
        print(f"❌ Ошибка HTTP: {e}")
        print(f"Статус: {response.status_code}, Ответ: {response.text}")
        try:
            error_details = response.json()
        except Exception:
            error_details = response.text
        return {"error": f"HTTP {response.status_code}", "details": error_details}
    except Exception as e:
        print(f"❌ Неизвестная ошибка: {e}")
        return {"error": str(e)}


def main():
    # import_data(DEAL_ID)
    # export_data(DEAL_ID)
    generate_3kp(DEAL_ID)


# === Запуск ===

def fill_deal_sheet(worksheet, deal_data, deal_id, start_row=2):
    """
    Заполняет вкладку "Сделка" основными полями и значениями из сделки
    Колонка A - название поля (как в Битриксе), колонка B - значение
    """
    # print("Заполняем вкладку 'Сделка'...")
    # print(f"deal_data keys: {list(deal_data.keys())[:10]}...")  # Показываем первые 10 ключей

    current_row = start_row

    # Создаем словарь соответствия: поле -> название поля (для отображения)
    field_names = {
        # 'ID': 'Номер сделки',
        # 'TITLE': 'Название сделки',
        # Добавляем пользовательские поля
        'UF_CRM_TP_1': 'Название закупки',
        'UF_CRM_TP_2': 'Цена контракта',
        'UF_CRM_TP_3': 'Дата начала подачи заявок',
        'UF_CRM_TP_4': 'Дата окончания подачи заявок',
        'UF_CRM_TP_5': 'Обеспечение заявки',
        'UF_CRM_TP_6': 'Обеспечение контракта',
        'UF_CRM_TP_MARKS': 'Метки',
        'UF_CRM_TP_7': 'Посмотреть в Тендерплане',
        'UF_CRM_TP_8': 'Тип размещения',
        'UF_CRM_TP_9': 'Закон/Площадка',
        'UF_CRM_TP_10': 'Посмотреть на площадке',
        'UF_CRM_TP_11': 'Номер тендера',
        'UF_CRM_TP_12': 'Размер обеспечения гарантийных обязательств',
        'UF_CRM_TP_13': 'Дата рассмотрения и оценки заявок',
        'UF_CRM_TP_14': 'Дата проведения аукциона',
        'UF_CRM_TP_15': 'Дата подведения итогов',
        'UF_CRM_TP_16': 'ИНН заказчика',
        'UF_CRM_1757398708466': 'Требуется монтаж/пусконаладочные работы',
        'UF_CRM_1757398736633': 'Требуется Обучение сотрудников',
        'UF_CRM_1757398787387': 'Требуется представитель от компании',
        'UF_CRM_1757398866815': 'Количество позиций:',
        'UF_CRM_1757398927169': 'Срок поставки товара:',
        'UF_CRM_1757402422471': 'Ограничение для иностранного товара',
        'UF_CRM_1757402456572': 'Преимущество российского товара',
        'UF_CRM_1757912892451': 'Причина отказа',
        'UF_CRM_1757916728790': 'Выбрал поставщика',
        'UF_CRM_1757929901175': 'Аналоги',
        'UF_CRM_1757930626746': 'Адрес доставки',
        'UF_CRM_1757931573446': 'Адрес отгрузки',
        'UF_CRM_1757999862739': 'Дата планируемой поставки',
        'UF_CRM_1759603831093': 'Компания отгрузки',
        'UF_CRM_1761537686'   : 'Наценка  итоговая'

    }

    # Специальное поле: ссылка на сделку
    deal_link = f'https://labkabinet.bitrix24.ru/crm/deal/details/{deal_id}/'

    # Заполняем все поля в порядке их следования в таблице
    for field_name in ['ID', 'TITLE'] + list(field_names.keys()):
        if field_name == 'TITLE':
            # Записываем название поля и его значение
            worksheet[f'B{current_row}'] = 'Название сделки'
            value = deal_data.get(field_name, '')
            worksheet[f'C{current_row}'] = value
            current_row += 1

            # Добавляем ссылку на сделку (специальное поле)
            worksheet[f'B{current_row}'] = 'Ссылка на сделку'
            worksheet[f'C{current_row}'] = deal_link
            current_row += 1
        elif field_name in deal_data:
            # Записываем название поля и его значение
            display_name = field_names.get(field_name, field_name)  # Если нет в словаре, используем имя поля
            worksheet[f'B{current_row}'] = display_name
            value = deal_data.get(field_name, '')
            if isinstance(value, (dict, list)):
                worksheet[f'C{current_row}'] = str(value)
            else:
                worksheet[f'C{current_row}'] = value
            current_row += 1

    # print(f"Заполнение вкладки 'Сделка' завершено. Записано {current_row - start_row} строк.")


def products_to_excel(deal_id, output_file=None, catalog_id=None):
    """
    Возвращает DataFrame с товарами сделки,
    колонки — в порядке COLUMN_LABELS,
    заголовки — русские метки из COLUMN_LABELS.
    """
    print(f"Загружаем товарные позиции для сделки {deal_id}...")

    deal_products = get_deal_products(deal_id)
    if not deal_products:
        print("В сделке нет товаров.")
        # Создаём пустой DataFrame с нужными колонками и русскими заголовками
        return pd.DataFrame(columns=list(COLUMN_LABELS.values()))

    all_rows = []

    for i, product_row in enumerate(deal_products):
        product_id = product_row.get('PRODUCT_ID')
        print(f"Обрабатываем позицию {i + 1}/{len(deal_products)} (ID: {product_id})")

        catalog_data = get_catalog_element(product_id) if product_id else None

        # Собираем плоский словарь со всеми возможными полями
        row_data = {}

        # Поля из строки сделки (без префикса DEAL_ROW_)
        for k, v in product_row.items():
            row_data[k] = v

        # Поля из каталога (если есть)
        if catalog_data:
            for k, v in catalog_data.items():
                # Не перезаписываем, если уже есть (например, PRICE из сделки важнее)
                if k not in row_data:
                    row_data[k] = v

            # Раскрываем PROPERTY_XXX
            for key in list(row_data.keys()):
                if key.startswith('PROPERTY_') and isinstance(row_data[key], dict):
                    row_data[key] = row_data[key].get('value', '')

        # Служебные поля
        # row_data['ID строки в сделке'] = product_row.get('ID')
        # row_data['Из каталога?'] = catalog_data is not None

        all_rows.append(row_data)

    # Создаём DataFrame с колонками в порядке COLUMN_LABELS (технические имена)
    technical_columns = list(COLUMN_LABELS.keys())
    df = pd.DataFrame(all_rows, columns=technical_columns)

    # Переименовываем колонки на русские метки (в том же порядке!)
    df.rename(columns=COLUMN_LABELS, inplace=True)

    return df

def create_products_from_tovary(deal_id, webhook_url=WEBHOOK_URL, output_dir=OUTPUT_DIR):
    input_file = os.path.join(output_dir, str(deal_id), f"расчет_{deal_id}.xlsx")
    if not os.path.exists(input_file):
        print(f"❌ Файл не найден: {input_file}")
        return

    wb = load_workbook(input_file)  # Без data_only — чтобы можно было писать обратно
    if "Товары" not in wb.sheetnames:
        print("❌ Лист 'Товары' не найден.")
        return

    ws = wb["Товары"]
    # Заголовки — технические имена из строки 1
    headers = []
    col = 1
    while True:
        val = ws.cell(row=1, column=col).value
        if val is None:
            break
        headers.append(str(val).strip())
        col += 1

    if not headers:
        print("❌ Не найдены заголовки на листе 'Товары'.")
        return

    try:
        id_col_idx = headers.index("PRODUCT_ID") + 1
        name_col_idx = headers.index("PRODUCT_NAME") + 1
    except ValueError as e:
        print(f"❌ Отсутствует обязательный столбец: {e}")
        return

    row_idx = 3
    created_count = 0
    while True:
        name_val = ws.cell(row=row_idx, column=name_col_idx).value
        if name_val is None or str(name_val).strip() == "":
            break

        product_id_val = ws.cell(row=row_idx, column=id_col_idx).value
        is_manual = (
            product_id_val is None
            or str(product_id_val).strip() == ""
            or (isinstance(product_id_val, (int, float)) and float(product_id_val) == 0)
        )

        if is_manual:
            # Собираем данные строки
            row_data = {}
            for i, key in enumerate(headers):
                val = ws.cell(row=row_idx, column=i + 1).value
                if val is not None:
                    row_data[key] = val

            name = str(row_data.get("PRODUCT_NAME", f"Товар {deal_id}-{row_idx}")).strip()
            if name.startswith("="):
                print(f"⚠️ Пропуск строки {row_idx}: название содержит формулу")
                row_idx += 1
                continue

            fields = {"NAME": name}
            if row_data.get("PRICE") not in (None, ""):
                try:
                    fields["PRICE"] = float(row_data["PRICE"])
                except:
                    pass
            if row_data.get("CURRENCY_ID"):
                fields["CURRENCY_ID"] = str(row_data["CURRENCY_ID"])
            if row_data.get("VAT_INCLUDED") in ("Y", "N"):
                fields["VAT_INCLUDED"] = row_data["VAT_INCLUDED"]

            # PROPERTY_XXX
            for key, val in row_data.items():
                if key.startswith("PROPERTY_") and key != "PROPERTY_202" and val not in (None, ""):
                    fields[key] = {"value": val}

            # Создаём товар
            try:
                resp = requests.post(f"{webhook_url}crm.product.add", json={"fields": fields})
                resp.raise_for_status()
                new_id = resp.json().get("result")
                if new_id:
                    ws.cell(row=row_idx, column=id_col_idx, value=new_id)
                    created_count += 1
                    print(f"✅ Создан товар ID={new_id} для строки {row_idx}")
                else:
                    print(f"⚠️ Не удалось получить ID для строки {row_idx}")
            except Exception as e:
                print(f"❌ Ошибка создания товара в строке {row_idx}: {e}")

        row_idx += 1

    if created_count > 0:
        wb.save(input_file)
        print(f"💾 Excel обновлён: добавлено {created_count} новых PRODUCT_ID.")
    else:
        print("ℹ️ Новых товаров для создания не найдено.")
        

if __name__ == "__main__":
    deal_id=13968
    #import_data(deal_id)
    #create_products_from_tovary(DEAL_ID)
    generate_kp_lsho(deal_id)
    # fill_excel(DEAL_ID)
    # deal_to_exel(DEAL_ID,deal)
    # auto_update_check()
    # import_data(DEAL_ID)
    #export_data(deal_id)
    # generate_3kp(DEAL_ID)
    # main()
