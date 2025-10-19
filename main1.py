import os
import shutil
import requests
from numba.core.typing.builtins import Print
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import json  # Импортируем json для сериализации сложных объектов
new_row["ставка НДС/входящий"]=20
# === Настройки ===
WEBHOOK_URL = "https://labkabinet.bitrix24.ru/rest/6808/9wti8nc7t0j9r2c7/"
DEAL_ID = 25034
INPUT_DIR="Шаблоны"
OUTPUT_DIR = "Расчеты"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"{DEAL_ID}\расчет_{DEAL_ID}.xlsx")
TEMPLATE_FILE = os.path.join(INPUT_DIR, f"Расчет_шаблон_V1.xlsx")

SUPPLIER_MAP = {
    "ООО": {
        "name": 'ООО "Научные развлечения"',
        "vat_in": "НДС 20%",
        "shipping_city": "Москва"
    },
    # Можно добавить другие поставщики
}

print("Запуск лога")
# === 1. Получение строк товаров из сделки ===
def get_deal_product(deal_id):
    url = f"{WEBHOOK_URL}crm.deal.productrows.get"
    response = requests.post(url, json={"id": deal_id})
    response.raise_for_status()
    return response.json().get("result", [])


# === 2. Получение данных товара из каталога магазина ===
def get_catalog_product(product_id):
    url = f"{WEBHOOK_URL}catalog.product.get"
    response = requests.post(url, json={"id": product_id})
    if response.status_code == 200:
        return response.json().get("result", {})
    else:
        print(f"⚠️ Не удалось загрузить товар ID={product_id}")
        return None





# === 3. Полная выгрузка товаров в Excel с динамическими полями и обработкой сложных типов ===
def export_products_to_db(deal_products, output_path):
    """
    Экспортирует данные о товарах из сделки в Excel файл с динамическими столбцами.
    Столбцы определяются на основе всех уникальных ключей из данных товаров (как из сделки, так и из каталога).
    Сложные типы данных (dict, list) сериализуются в строки JSON.
    :param deal_products: Список словарей, представляющих строки товаров из сделки (результат get_deal_product).
    :param output_path: Путь к Excel файлу для сохранения.
    """
    # Список для хранения всех обработанных строк (с объединенными данными)
    all_rows_data = []
    # Множество для хранения всех уникальных ключей (названий столбцов)
    all_keys = set()

    print("Обработка товаров из сделки...")
    for product_row in deal_products:
        catalog_product_id = product_row.get("PRODUCT_ID")

        catalog_data = {}
        if catalog_product_id:
            catalog_data = get_catalog_product(catalog_product_id) or {}

        combined_row = {}
        # Добавляем префиксы
        for key, value in product_row.items():
            combined_row[f"DEAL_{key}"] = value

        for key, value in catalog_data.items():
            combined_row[f"CATALOG_{key}"] = value

        # Обработка значений перед добавлением в combined_row
        processed_row = {}
        for key, original_value in combined_row.items():
            # Проверяем тип значения
            if isinstance(original_value, (dict, list)):
                # Сериализуем сложный объект в строку JSON
                processed_value = json.dumps(original_value, ensure_ascii=False,
                                             indent=2)  # ensure_ascii=False для кириллицы
            else:
                # Оставляем примитивные типы как есть, или конвертируем в строку, если нужно
                # Если вы хотите все значения в виде строк, используйте str(original_value)
                # processed_value = str(original_value)
                # Но лучше оставить как есть, если это примитив, и только сложные типы сериализовать
                processed_value = original_value

            processed_row[key] = processed_value

        all_rows_data.append(processed_row)
        all_keys.update(processed_row.keys())

    headers = sorted(list(all_keys))

    print(f"Найдено уникальных полей (столбцов): {len(headers)}")
    if not headers:
        print("Предупреждение: Не найдено никаких данных для экспорта.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Товары_Сделки"

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    for row_num, row_data in enumerate(all_rows_data, 2):
        for col_num, header in enumerate(headers, 1):
            value = row_data.get(header, "")  # Значение уже обработано в processed_row
            ws.cell(row=row_num, column=col_num, value=value)

    # Автоподгонка ширины столбцов (может быть медленно)
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                # Для строк, полученных из JSON, длина может быть большой
                # Рассмотрите ограничение длины, если столбцы получаются слишком широкими
                str_val = str(cell.value)
                # Ограничиваем длину для расчета ширины, если значение слишком длинное
                display_val = str_val[:50] + "..." if len(str_val) > 50 else str_val
                if len(display_val) > max_length:
                    max_length = len(display_val)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Ограничиваем максимальную ширину
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"Полная выгрузка товаров успешно экспортирована в {output_path}")
    print(f"Файл содержит {len(all_rows_data)} строк товаров и {len(headers)} столбцов.")


def fill_excel(products, output_path):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    shutil.copy(TEMPLATE_FILE, output_path)

    wb = load_workbook(output_path)
    ws_calc = wb["Калькулятор"]
    ws_ship = wb["Доставка"]

    # === 1. Лист "Доставка" — заполняем только входные данные ===
    # Очищаем только входные столбцы (A–H), формулы в I–K останутся
    for row in range(3, 43):
        for col in "ABCDEFGH":
            ws_ship[f"{col}{row}"].value = None

    for i, p in enumerate(products):
        r = i + 3
        ws_ship[f"B{r}"] = p["name"]
        ws_ship[f"C{r}"] = p["quantity"]
        ws_ship[f"D{r}"] = p["supplier"]
        ws_ship[f"E{r}"] = p["shipping_city"]
        ws_ship[f"F{r}"] = p["weight_g"]
        ws_ship[f"G{r}"] = p["length_mm"]
        ws_ship[f"H{r}"] = p["width_mm"]
        ws_ship[f"I{r}"] = p["height_mm"]

    # Строка "Итого" на листе "Доставка" — формулы уже есть, только текст
    last_row_ship = len(products) + 2
    ws_ship[f"A{last_row_ship}"] = "Итого"

    # === 2. Лист "Калькулятор" — заполняем ТОЛЬКО входные ячейки ===
    start_row = 3
    max_rows = 40  # строки 3–42

    # Очищаем только входные столбцы (B–G), остальное — формулы!
    for i in range(max_rows):
        r = start_row + i
        for col in "BCDEFG":
            ws_calc[f"{col}{r}"].value = None

    # Заполняем реальные товары
    for i, p in enumerate(products):
        r = start_row + i
        ws_calc[f"B{r}"] = p["product_id"]
        ws_calc[f"C{r}"] = p["name"]
        ws_calc[f"D{r}"] = p["quantity"]
        ws_calc[f"E{r}"] = p["price_with_vat"]
        ws_calc[f"F{r}"] = p["supplier"]
        ws_calc[f"G{r}"] = p["vat_in"]
        ws_calc[f"H{r}"] = p["link"]

    # Строка "Доставка" (43) — заполняем только входные поля
    #ws_calc[f"С43"] = "Доставка"
    #ws_calc[f"D43"] = 1
    #ws_calc[f"D43"] = 27900  # можно параметризовать
    #ws_calc[f"E43"] = "СДЭК"
    #ws_calc[f"F43"] = "УСН"
    #ws_calc[f"G43"] = ""

    # Строка "Итого" (44) — только текст, формулы уже есть
    ws_calc[f"B44"] = "Итого"

    # === Скрываем пустые строки с товарами ===
    num_products = len(products)
    for i in range(num_products, max_rows):
        r = start_row + i
        ws_calc.row_dimensions[r].hidden = True

    # Строки "Доставка" (43) и "Итого" (44) — всегда видимы
    ws_calc.row_dimensions[43].hidden = False
    ws_calc.row_dimensions[44].hidden = False

    wb.save(output_path)
    return output_path

def import_data (deal_id):
    print("1️⃣ Получаем строки товаров из сделки...")
    rows = get_deal_product(deal_id)
    if not rows:
        print("❌ В сделке нет товаров.")
        return

    print("2️⃣ Загружаем данные из каталога...")
    products_for_excel = []

    for row in rows:
        product_id = row.get("PRODUCT_ID")
        name = row.get("PRODUCT_NAME", "").strip()
        quantity = row.get("QUANTITY", 1)
        price_with_vat = row.get("PRICE", 0)

        # Если название пустое — пропускаем или подставляем заглушку
        if not name:
            name = f"Товар ID={product_id}" if product_id else "Неизвестный товар"

        catalog_data = None
        if product_id:
            catalog_data = get_catalog_product(product_id)

        # Извлечение данных из каталога
        product = catalog_data.get("product", {}) if catalog_data else {}

        # Поставщик
        supplier_enum = product.get("property196", {}).get("valueEnum", "")
        supplier_info = SUPPLIER_MAP.get(supplier_enum, {
            "name": supplier_enum or "Не указан",
            "vat_in": "НДС 20%",
            "shipping_city": "Москва"
        })

        # Габариты и вес
        weight_g = product.get("weight", 0)
        length_mm = product.get("length", 0)
        width_mm = product.get("width", 0)
        height_mm = product.get("height", 0)

        # Ссылка — пока не приходит, но можно добавить позже
        link = ""

        products_for_excel.append({
            "product_id": product_id,
            "name": name,
            "quantity": quantity,
            "price_with_vat": price_with_vat,
            "supplier": supplier_info["name"],
            "vat_in": supplier_info["vat_in"],
            "shipping_city": supplier_info["shipping_city"],
            "weight_g": weight_g,
            "length_mm": length_mm,
            "width_mm": width_mm,
            "height_mm": height_mm,
            "link": link,
        })

    print("3️⃣ Заполняем Excel...")
    output_file = fill_excel(products_for_excel, OUTPUT_FILE)
    print(f"✅ Готово! Файл сохранён: {output_file}")


def export_data(deal_id):
    print(f"🔧 Экспорт данных в сделку {deal_id}")
    input_file = os.path.join(OUTPUT_DIR, str(deal_id), f"расчет_{deal_id}.xlsx")
    if not os.path.exists(input_file):
        print(f"❌ Файл не найден: {input_file}")
        return

    wb = load_workbook(input_file, data_only=True)
    ws_calc = wb["Калькулятор"]

    new_rows = []
    for row in range(3, 43):
        product_id_cell = ws_calc[f"B{row}"].value
        if product_id_cell in (None, ""):
            # Продолжаем, пока не встретим пустую строку ИЛИ дойдём до "Доставка"
            name = ws_calc[f"C{row}"].value
            if name and str(name).strip().lower() in ("доставка", "итого"):
                break
            if not name:
                break

        try:
            product_id = int(float(product_id_cell)) if product_id_cell not in (None, "") else 0
        except (ValueError, TypeError):
            product_id = 0

        name = str(ws_calc[f"C{row}"].value or "").strip()
        quantity = ws_calc[f"D{row}"].value or 1
        price = ws_calc[f"E{row}"].value or 0

        # Пропускаем полностью пустые строки
        if not name and product_id == 0 and quantity == 1 and price == 0:
            continue

        try:
            quantity = int(float(quantity))
            price = float(price)
        except:
            continue

        if product_id == 0:
            # Ручной товар — без PRODUCT_ID
            if name:  # только если есть название
                new_rows.append({
                    "PRODUCT_NAME": name,
                    "QUANTITY": quantity,
                    "PRICE": price,
                })
                print(f"✏️ Ручной товар: {name} × {quantity} = {price}")
        else:
            # Товар из каталога
            new_rows.append({
                "PRODUCT_ID": product_id,
                "QUANTITY": quantity,
                "PRICE": price,
            })
            print(f"📦 Каталожный товар ID={product_id}: {name} × {quantity} = {price}")

    if not new_rows:
        print("⚠️ Нет строк для обновления.")
        return

    print(f"\n📤 Отправка {len(new_rows)} строк в сделку {deal_id}...")
    payload = {"id": deal_id, "rows": new_rows}
    print("Тело запроса (первые 3 строки):")
    for i, r in enumerate(new_rows[:3]):
        print(f"  {i+1}. {r}")

    try:
        response = requests.post(
            f"{WEBHOOK_URL}crm.deal.productrows.set",
            json=payload
        )
        response.raise_for_status()
        print("✅ Сделка успешно обновлена!")
    except Exception as e:
        print(f"❌ Ошибка: {e}")
        if 'response' in locals():
            print("Ответ Bitrix24:", response.text)
        return

    print("✅ Экспорт завершён.")
# === Основная логика ===

def export_data_КЕДО(deal_id):
    print("\n📤 ЭКСПОРТ КЕДО (цены из колонки x = 24)")
    _export_data_with_price_column(
        deal_id,
        price_col_index=24,          # колонка х
        tax_rate="0.00",             # Без НДС
        tax_included="N",            # НДС не включён в цену
        supplier_name="ИП",
        mode="КЕДО"
    )


def export_data_Verch(deal_id):
    print("\n📤 ЭКСПОРТ ВЕРШ (цены из колонки W = 23)")
    _export_data_with_price_column(
        deal_id,
        price_col_index=23,          # колонка T
        tax_rate="0.00",             # Без НДС
        tax_included="N",            # НДС не включён
        supplier_name="ИП",
        mode="Верш"
    )


def export_data_LSO(deal_id):
    print("\n📤 ЭКСПОРТ ЛШО (цены из колонки V = 22)")
    _export_data_with_price_column(
        deal_id,
        price_col_index=22,          # колонка Q — "Цена ЛШО"
        tax_rate="20.00",            # НДС 20%
        tax_included="Y",            # НДС включён в цену
        supplier_name="ИП",
        mode="ЛШО"
    )

def _export_data_with_price_column(deal_id, price_col_index, tax_rate, tax_included, supplier_name, mode):
    input_file = os.path.join(OUTPUT_DIR, str(deal_id), f"расчет_{deal_id}.xlsx")
    if not os.path.exists(input_file):
        print(f"❌ Файл не найден: {input_file}")
        return

    wb = load_workbook(input_file, data_only=True)
    ws = wb["Калькулятор"]

    new_rows = []
    for row in range(3, 43):
        product_id_raw = ws.cell(row=row, column=2).value  # B
        name = str(ws.cell(row=row, column=3).value or "").strip()

        if not name or name.lower() in ("доставка", "итого"):
            if product_id_raw in (None, "") and not name:
                break
            # "Доставка" и "Итого" не добавляем как товары

        try:
            product_id = int(float(product_id_raw)) if product_id_raw not in (None, "") else 0
        except (ValueError, TypeError):
            product_id = 0

        quantity_raw = ws.cell(row=row, column=4).value  # D
        price_raw = ws.cell(row=row, column=price_col_index).value  # нужная колонка

        quantity = int(float(quantity_raw)) if quantity_raw not in (None, "") else 1
        price = float(price_raw) if price_raw not in (None, "") else 0.0

        row_data = {
            "QUANTITY": quantity,
            "PRICE": price,
            "TAX_RATE": tax_rate,
            "TAX_INCLUDED": tax_included,
        }

        if product_id == 0:
            if name:
                row_data["PRODUCT_NAME"] = name
                new_rows.append(row_data)
        else:
            row_data["PRODUCT_ID"] = product_id
            new_rows.append(row_data)

        print(f"   ✅ {name} → {price} руб, НДС: {tax_rate}, Включён: {tax_included}")

    # Отправка в Bitrix24
    try:
        response = requests.post(
            f"{WEBHOOK_URL}crm.deal.productrows.set",
            json={"id": deal_id, "rows": new_rows}
        )
        response.raise_for_status()
        print(f"✅ Успешно обновлено ({mode})!")
    except Exception as e:
        print(f"❌ Ошибка при экспорте {mode}: {e}")
        if 'response' in locals():
            print(f"Ответ Bitrix24:", json.dumps(response.text))

def generate_3kp(deal_id):
    """
    Генерация 3 КП:
    - КЕДО (цены из колонки U)
    - Верш (цены из колонки T)
    """
    print(f"🖨️ Генерация 3 КП для сделки {deal_id}...")

    # Экспорт для КЕДО
    print("\n➡️ Экспорт КЕДО...")
    export_data_КЕДО(DEAL_ID)
    generate_KP(DEAL_ID, 50)
    print("\n➡️ Экспорт Верхозина...")
    export_data_Verch(DEAL_ID)
    generate_KP(DEAL_ID, 48)
    print("\n➡️ Экспорт ЛШО...")
    export_data_LSO(DEAL_ID)
    generate_KP(DEAL_ID, 46)


    print("\n✅ Генерация 3 КП завершена.")



def generate_KP(entity_id, template_id=46, webhook_url=WEBHOOK_URL, entity_type_id=2, output_dir="Расчеты"):
    """
    Вызывает метод crm.documentgenerator.document.add через вебхук,
    имитируя BX24.callMethod, и сохраняет сгенерированный документ в указанный каталог.

    Args:
        entity_id (str or int): ID сущности (например, сделки).
        template_id (int): ID шаблона:
            46 → "КП ЛШО",
            48 → "КП Верхозина",
            50 → "КП КЕДО",
            52 → "КП ЛШО с фото".
        webhook_url (str): URL вебхука Bitrix24.
        entity_type_id (int or str): Тип сущности (2 = сделка).
        output_dir (str): Директория для сохранения файлов.

    Returns:
        dict: Результат API + путь к сохранённому файлу (если успешно).
    """

    # Определяем префикс имени файла по template_id
    template_names = {
        46: "КП ЛШО",
        48: "КП Верхозина",
        50: "КП КЕДО",
        52: "КП ЛШО с фото"
    }
    template_name = template_names.get(template_id, f"КП_шаблон_{template_id}")

    # Формируем путь к файлу
    filename = f"{template_name}_{entity_id}.docx"
    deal_subdir = os.path.join(output_dir, str(entity_id))
    full_path = os.path.join(deal_subdir, filename)

    def download_document(download_url, save_path):
        """Загружает документ и сохраняет по указанному пути."""
        if not download_url:
            print("❌ Отсутствует downloadUrl для загрузки документа.")
            return None

        try:
            os.makedirs(os.path.dirname(save_path), exist_ok=True)
            print(f"📥 Загрузка документа по URL: {download_url}")
            resp = requests.get(download_url)
            resp.raise_for_status()

            with open(save_path, 'wb') as f:
                f.write(resp.content)
            print(f"✅ Документ успешно сохранён: {save_path}")
            return save_path

        except Exception as e:
            print(f"❌ Ошибка при загрузке документа: {e}")
            return None

    # --- Основной вызов API ---
    api_method = 'crm.documentgenerator.document.add'
    url = f"{webhook_url.rstrip('/')}/{api_method}.json"

    payload = {
        'templateId': template_id,
        'entityTypeId': entity_type_id,
        'entityId': entity_id
    }

    print(f"Попытка вызвать метод {api_method} с параметрами: {payload}")
    print(f"URL запроса: {url}")

    try:
        response = requests.post(url, json=payload)
        response.raise_for_status()
        result = response.json()

        print("--- Результат вызова API ---")
        print(json.dumps(result, indent=2, ensure_ascii=False))

        if 'result' in result and 'document' in result['result']:
            download_url = result['result']['document'].get('downloadUrlMachine')
            if download_url:
                saved_path = download_document(download_url, full_path)
                if saved_path:
                    result['downloaded_file'] = saved_path
            else:
                print("⚠️ В ответе отсутствует downloadUrl.")
            print("Документ успешно создан.")
            return result
        else:
            print(f"⚠️ Некорректный ответ API: {result}")
            return result

    except requests.exceptions.HTTPError as e:
        print(f"❌ Ошибка HTTP: {e}")
        print(f"Статус: {response.status_code}, Ответ: {response.text}")
        try:
            error_details = response.json()
        except Exception:
            error_details = response.text
        return {"error": f"HTTP {response.status_code}", "details": error_details}
    except Exception as e:
        print(f"❌ Неизвестная ошибка: {e}")
        return {"error": str(e)}

def main():
    #import_data(DEAL_ID)
    #export_data(DEAL_ID)
    generate_3kp(DEAL_ID)



# === Запуск ===
if __name__ == "__main__":
    #import_data(DEAL_ID)
    #export_data(DEAL_ID)
    generate_3kp(DEAL_ID)
    #main()

